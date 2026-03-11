import customtkinter as ctk
from tkinter import messagebox, filedialog
from theme import *
from database import get_assets_for_report, get_all_categories
import os, subprocess, sys
from datetime import datetime


class ReportsPage(ctk.CTkFrame):
    def __init__(self, parent, navigate, current_user):
        super().__init__(parent, fg_color="transparent")
        self.navigate = navigate
        self.current_user = current_user
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._cat_map = {}
        self._build()

    def _build(self):
        # Header
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(hdr, text="Reports", font=FONT_TITLE(), text_color=TEXT_MAIN).pack(anchor="w")
        ctk.CTkLabel(hdr, text="Export inventory data to Excel spreadsheets.",
                     font=FONT_BODY(), text_color=TEXT_DIM).pack(anchor="w")

        # Report cards
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent", corner_radius=0)
        scroll.grid(row=1, column=0, sticky="nsew")
        scroll.grid_columnconfigure(0, weight=1)

        # ── Full Inventory Report ──────────────────────────────────────────────
        self._report_card(
            scroll,
            icon="📊",
            title="Full Inventory Report",
            description="Export all assets with their category, supplier, quantity, value, condition and location.",
            color=ACCENT,
            build_fn=self._export_full
        )

        # ── Low Stock Report ───────────────────────────────────────────────────
        self._report_card(
            scroll,
            icon="⚠️",
            title="Low Stock Report",
            description="Export only assets that are at or below their minimum stock level.",
            color=WARN,
            build_fn=self._export_low_stock
        )

        # ── Category Report ────────────────────────────────────────────────────
        cat_card = ctk.CTkFrame(scroll, fg_color=BG_CARD, corner_radius=14)
        cat_card.pack(fill="x", pady=8)
        cat_card.grid_columnconfigure(1, weight=1)

        icon_bg = ctk.CTkFrame(cat_card, fg_color="#a78bfa", corner_radius=12, width=56, height=56)
        icon_bg.grid(row=0, column=0, rowspan=2, padx=16, pady=16)
        icon_bg.grid_propagate(False)
        ctk.CTkLabel(icon_bg, text="📁", font=("Segoe UI Emoji", 24)).place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(cat_card, text="Category Report", font=FONT_BOLD(15),
                     text_color=TEXT_MAIN).grid(row=0, column=1, sticky="w", pady=(16, 2))
        ctk.CTkLabel(cat_card, text="Export assets filtered by a specific category.",
                     font=FONT_SMALL(), text_color=TEXT_DIM).grid(row=1, column=1, sticky="w", pady=(0, 4))

        opts_frame = ctk.CTkFrame(cat_card, fg_color="transparent")
        opts_frame.grid(row=2, column=1, sticky="ew", pady=(4, 16))

        cats = get_all_categories()
        self._cat_map = {c["name"]: c["id"] for c in cats}
        cat_names = [c["name"] for c in cats] or ["No categories"]

        self.dd_cat = ctk.CTkOptionMenu(opts_frame, values=cat_names, width=200, height=34,
                                        fg_color=BG_INPUT, button_color=BG_INPUT,
                                        button_hover_color=BORDER, text_color=TEXT_MAIN,
                                        font=FONT_BODY())
        self.dd_cat.pack(side="left", padx=(0, 10))
        ctk.CTkButton(opts_frame, text="Export", width=100, height=34, corner_radius=8,
                      fg_color="#a78bfa", hover_color="#8b5cf6",
                      command=self._export_by_category).pack(side="left")

        # ── Summary Stats Card ─────────────────────────────────────────────────
        self._report_card(
            scroll,
            icon="📋",
            title="Summary Report",
            description="A summary sheet with totals: asset count, total value, categories, and low stock count.",
            color=ACCENT2,
            build_fn=self._export_summary
        )

    def _report_card(self, parent, icon, title, description, color, build_fn):
        card = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=14)
        card.pack(fill="x", pady=8)
        card.grid_columnconfigure(1, weight=1)

        icon_bg = ctk.CTkFrame(card, fg_color=color, corner_radius=12, width=56, height=56)
        icon_bg.grid(row=0, column=0, rowspan=2, padx=16, pady=16)
        icon_bg.grid_propagate(False)
        ctk.CTkLabel(icon_bg, text=icon, font=("Segoe UI Emoji", 24)).place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(card, text=title, font=FONT_BOLD(15),
                     text_color=TEXT_MAIN).grid(row=0, column=1, sticky="w", pady=(16, 2))
        ctk.CTkLabel(card, text=description, font=FONT_SMALL(),
                     text_color=TEXT_DIM).grid(row=1, column=1, sticky="w", pady=(0, 16))

        ctk.CTkButton(card, text="Export to Excel", width=140, height=36, corner_radius=8,
                      fg_color=color, hover_color=BG_CARD,
                      command=build_fn).grid(row=0, column=2, rowspan=2, padx=16)

    # ── Export helpers ─────────────────────────────────────────────────────────
    def _pick_save_path(self, default_name):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=default_name
        )
        return path or None

    def _open_file(self, path):
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass

    def _export_full(self):
        self._do_export(get_assets_for_report(), "AssetMate_Full_Inventory")

    def _export_low_stock(self):
        from database import get_low_stock_assets
        self._do_export(get_low_stock_assets(), "AssetMate_LowStock")

    def _export_by_category(self):
        cat_name = self.dd_cat.get()
        cat_id   = self._cat_map.get(cat_name)
        if not cat_id:
            messagebox.showwarning("No Category", "Please select a valid category.")
            return
        assets = get_assets_for_report(category_id=cat_id)
        safe_name = cat_name.replace(" ", "_").replace("/", "-")
        self._do_export(assets, f"AssetMate_{safe_name}")

    def _export_summary(self):
        from database import get_dashboard_stats
        stats = get_dashboard_stats()
        self._do_summary_export(stats)

    def _do_export(self, assets, base_name):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            if messagebox.askyesno("Install Required",
                                   "openpyxl is needed for Excel export.\nInstall it now?"):
                import subprocess, sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            else:
                return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self._pick_save_path(f"{base_name}_{timestamp}.xlsx")
        if not path: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Styles
        header_fill  = PatternFill("solid", fgColor="1e2535")
        header_font  = Font(bold=True, color="4f8ef7", size=11)
        title_font   = Font(bold=True, color="e8eaf0", size=13)
        border_side  = Side(style="thin", color="2a3347")
        thin_border  = Border(left=border_side, right=border_side,
                              top=border_side, bottom=border_side)
        center       = Alignment(horizontal="center", vertical="center")
        alt_fill     = PatternFill("solid", fgColor="252d3d")
        warn_fill    = PatternFill("solid", fgColor="3d2e10")
        warn_font    = Font(color="f5a623", size=10)

        # Title row
        ws.merge_cells("A1:J1")
        ws["A1"] = f"AssetMate — Inventory Report   |   Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws["A1"].font      = title_font
        ws["A1"].fill      = PatternFill("solid", fgColor="0f1117")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        # Headers
        headers = ["ID", "Name", "Category", "Supplier", "Qty", "Min Stock",
                   "Unit Value (₱)", "Total Value (₱)", "Condition", "Location"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border
        ws.row_dimensions[2].height = 22

        # Data rows
        for i, a in enumerate(assets, 3):
            low       = a.get("quantity", 0) <= a.get("min_stock", 1)
            row_fill  = warn_fill if low else (alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="1e2535"))
            row_font  = warn_font if low else Font(color="e8eaf0", size=10)
            total_val = (a.get("unit_value") or 0) * (a.get("quantity") or 0)
            values = [
                a.get("id"), a.get("name"),
                a.get("category_name") or "—",
                a.get("supplier_name") or "—",
                a.get("quantity"), a.get("min_stock"),
                a.get("unit_value"), round(total_val, 2),
                a.get("condition"), a.get("location") or "—"
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.fill      = row_fill
                cell.font      = row_font
                cell.border    = thin_border
                cell.alignment = center
            ws.row_dimensions[i].height = 18

        # Column widths
        col_widths = [6, 30, 18, 18, 8, 10, 16, 16, 14, 16]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        ws.freeze_panes = "A3"
        wb.save(path)
        messagebox.showinfo("✅ Export Complete", f"Saved to:\n{path}")
        self._open_file(path)

    def _do_summary_export(self, stats):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing", "Run: pip install openpyxl")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self._pick_save_path(f"AssetMate_Summary_{timestamp}.xlsx")
        if not path: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        title_font = Font(bold=True, color="4f8ef7", size=14)
        val_font   = Font(bold=True, color="e8eaf0", size=12)
        lbl_font   = Font(color="6b7a99", size=11)
        bg         = PatternFill("solid", fgColor="1e2535")
        center     = Alignment(horizontal="center")

        ws.merge_cells("A1:B1")
        ws["A1"] = "AssetMate — Summary Report"
        ws["A1"].font      = title_font
        ws["A1"].fill      = PatternFill("solid", fgColor="0f1117")
        ws["A1"].alignment = center

        rows = [
            ("Total Asset Types",  stats["total_assets"]),
            ("Total Units",        stats["total_units"]),
            ("Total Value (₱)",    f"₱{stats['total_value']:,.2f}"),
            ("Low Stock Items",    stats["low_stock_count"]),
            ("Categories",         stats["total_categories"]),
            ("Suppliers",          stats["total_suppliers"]),
            ("Report Date",        datetime.now().strftime("%B %d, %Y")),
        ]
        for i, (label, value) in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=label).font = lbl_font
            ws.cell(row=i, column=2, value=value).font = val_font
            ws.cell(row=i, column=1).fill = bg
            ws.cell(row=i, column=2).fill = bg

        wb.save(path)
        messagebox.showinfo("✅ Export Complete", f"Saved to:\n{path}")
        self._open_file(path)